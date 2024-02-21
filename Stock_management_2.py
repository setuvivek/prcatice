from openpyxl import load_workbook

xpath = '/home/setu/workspace/python/pythonProject/Riken/stock_management.xlsx'
ws = load_workbook(xpath)
if "cust" and "prod" and "stock" not in ws.sheetnames:
    ws.create_sheet("cust")
    ws["cust"].append(("cid", "cname", "ctype1", "cpan"))
    ws.create_sheet("prod")
    ws["prod"].append(("pid", "pname", "incoming", "outgoing", "onhand"))
    ws.create_sheet("stock")
    ws["stock"].append(("stock id", "cust id", "prod id", "quantity", "in_out"))
    ws.save(xpath)
    print(ws["cust"].max_row)
    print(ws["prod"].max_row)
    print(ws["stock"].max_row)

print(ws.sheetnames)
list1 = []
list2 = []
list3 = []


def input_id(message):
    while True:
        try:
            return input(message)
        except ValueError:
            print("Please enter an integer")


def check_duplicate(data_list, index, value):
    return any(i[index] == value for i in data_list)


class Customer:
    def __init__(self, cid, cname, ctype1, cpan):
        self.id = cid
        self.name = cname
        self.type1 = ctype1
        self.pan = cpan

    @staticmethod
    def customer():
        seet1 = ws["cust"]
        while True:
            cid = input_id("Enter Customer Id:")
            if check_duplicate(list1, 0, cid) or check_duplicate(ws["cust"].iter_rows(values_only=True), 0, cid):
                print("Customer already exists")
                return Customer.customer()

            cname = input("Enter Customer Name: ")
            ctype1 = str(input("Enter Customer Type customer/vendor "))
            cpan = str(input("Enter Customer Pan: "))
            if check_duplicate(list1, 3, cpan) or check_duplicate(ws["cust"].iter_rows(values_only=True), 3, cpan):
                print("Customer pan number must be unique.")
                return Customer.customer()

            list1.append([cid, cname, ctype1, cpan])
            seet1.append((cid, cname, ctype1, cpan))
            ws.save(xpath)
            main()


class Product:
    def __init__(self, pid, pname, incoming, outgoing, onhand):
        self.id = pid
        self.name = pname
        self.incoming = incoming
        self.outgoing = outgoing
        self.onhand = onhand

    @staticmethod
    def product():
        sheet2 = ws["prod"]
        while True:
            pid = input_id("Enter Product Id:")
            if check_duplicate(list2, 0, pid) or check_duplicate(ws["prod"].iter_rows(values_only=True), 0, pid):
                print("Product already exists")
                return Product.product()

            pname = str(input("Enter Product Name: "))
            if check_duplicate(list2, 1, pname) or check_duplicate(ws["prod"].iter_rows(values_only=True), 1, pname):
                print("Product with the same name already exists")
                return Product.product()

            incoming = int(input("Enter Incoming: "))
            outgoing = int(input("Enter Outgoing: "))
            if outgoing > incoming:
                print("Not valid")
                return Product.product()

            onhand = incoming - outgoing
            print(onhand)
            list2.append([pid, pname, incoming, outgoing, onhand])
            sheet2.append((pid, pname, incoming, outgoing, onhand))
            ws.save(xpath)
            main()


class Stock:
    def __init__(self, sid, spid, scid, sqty, stype2):
        self.id = sid
        self.product_id = spid
        self.customer_id = scid
        self.qty = sqty
        self.type2 = stype2

    @staticmethod
    def stock():
        sheet3 = ws["stock"]
        while True:
            sid = input_id("Enter Stock Id:")
            if check_duplicate(list3, 0, sid) or check_duplicate(ws["stock"].iter_rows(values_only=True), 0, sid):
                print("Stock entry already exists")
                return Stock.stock()

            spid = input("Enter Product Id:")
            if not check_duplicate(list2, 0, spid) and not check_duplicate(ws["prod"].iter_rows(values_only=True), 0,
                                                                           spid):
                print("This product does not exist.")
                return Stock.stock()

            scid = input("Enter Customer Id:")
            if not check_duplicate(list1, 0, scid) and not check_duplicate(ws["cust"].iter_rows(values_only=True), 0,
                                                                           scid):
                print("Customer does not exist.")
                return Stock.stock()


            if check_duplicate(ws["stock"].iter_rows(values_only=True), 0, sid) or \
                    check_duplicate(ws["stock"].iter_rows(values_only=True), 1, spid) or \
                    check_duplicate(ws["stock"].iter_rows(values_only=True), 2, scid):
                print("Stock entry with the same Stock ID, Product ID, or Customer ID already exists.")
                return Stock.stock()

            sqty = int(input("Enter qty:"))
            stype2 = str(input("Enter Type incoming/outgoing:"))

            if stype2 == "incoming":
                for i, item in enumerate(list2):
                    if item[0] == spid:
                        list2[i] = list(item)
                        list2[i][2] += sqty
                        list2[i][4] = list2[i][2] - list2[i][3]

                for row in ws["prod"].iter_rows(values_only=True):
                    if row[0] == spid:
                        new_row = list(row)
                        new_row[2] += sqty
                        new_row[4] = new_row[2] - new_row[3]
                        ws["prod"].append(new_row)

                list3.append([sid, spid, scid, sqty, stype2])
                sheet3.append((sid, spid, scid, sqty, stype2))

            elif stype2 == "outgoing":
                for i, item in enumerate(list2):
                    if item[0] == spid:
                        if sqty > item[4]:
                            print("Not enough available stock")
                            return Stock.stock()
                        else:
                            list2[i] = list(item)
                            list2[i][3] += sqty
                            list2[i][4] = list2[i][2] - list2[i][3]

                for row in ws["prod"].iter_rows(values_only=True):
                    if row[0] == spid:
                        if sqty > row[4]:
                            print("Not enough available stock")
                            return Stock.stock()
                        else:
                            new_row = list(row)
                            new_row[3] += sqty
                            new_row[4] = new_row[2] - new_row[3]
                            ws["prod"].append(new_row)

                list3.append([sid, spid, scid, sqty, stype2])
                sheet3.append((sid, spid, scid, sqty, stype2))

            ws.save(xpath)
            main()


def main():
    c = input("Enter which data you want to add customer, product or stock or exit: ")

    if c == "exit":
        print("Data of customer: ")
        for item in list1:
            print(item)
        print("Data of product: ")
        for item in list2:
            print(item)
        print("Data of stock: ")
        for item in list3:
            print(item)
        ws.close()
        exit()

    if c == "customer":
        Customer.customer()

    elif c == "product":
        Product.product()

    elif c == "stock":
        Stock.stock()

    return main()


if __name__ == "__main__":
    main()
