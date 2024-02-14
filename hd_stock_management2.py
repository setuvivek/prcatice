from openpyxl import Workbook, load_workbook

class Customer:
    def __init__(self):
        self.customer = []

    def add_customer(self, id, name, type, pan):
        if any(cust['ID'] == id for cust in self.customers):
            print("ID must be unique.")
            return
        if any(cust['PAN'] == pan for cust in self.customers):
            print("PAN must be unique.")
            return
        if type.lower() not in ['customer', 'vendor']:
            print("Customer type must be either 'customer' or 'vendor'.")
            return

        self.customers.append({'ID': id, 'Name': name, 'Type': type, 'PAN': pan})

    def save_to_excel(self, filename='customers.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Name', 'Type', 'PAN'])
        for cust in self.customers:
            ws.append([cust['ID'], cust['Name'], cust['Type'], cust['PAN']])
        wb.save(filename)

    def read_from_excel(self, filename='customers.xlsx'):
        try:
            wb = load_workbook(filename)
            ws = wb.active
            self.customers = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                id, name, type, pan = row
                self.customers.append({'ID': id, 'Name': name, 'Type': type, 'PAN': pan})
            print("Previous customer data:")
            for cust in self.customers:
                print(cust)
        except FileNotFoundError:
            print("No previous customer data found.")

class Product:
    def __init__(self):
        self.products = []

    def add_product(self, id, name, incoming, outgoing):
        if any(prod['ID'] == id for prod in self.products):
            print("ID must be unique.")
            return

        if any(prod['Name'] == name for prod in self.products):
            print("Name must be unique.")
            return

        if any(prod['Incoming'] < prod['Outgoing'] for prod in self.products):
            print("Stock is not available")
            return
        on_hand = incoming - outgoing

        if any(prod['OnHand'] < 0 for prod in self.products):
            print("OnHand is not negative")
            return

        # on_hand = incoming - outgoing
        self.products.append({'ID': id, 'Name': name, 'Incoming': incoming, 'Outgoing': outgoing, 'OnHand': on_hand})

    def update_stock(self, id, type, qty):
        for prod in self.products:
            if prod['ID'] == id:
                if type == 'incoming':
                    prod['Incoming'] += qty
                elif type == 'outgoing':
                    if prod['OnHand'] >= qty:
                        prod['Outgoing'] += qty
                    else:
                        print("Outgoing does not exists stock quantity")
                        return
                    # prod['Outgoing'] += qty
                prod['OnHand'] = prod['Incoming'] - prod['Outgoing']

    def save_to_excel(self, filename='products.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Name', 'Incoming', 'Outgoing', 'OnHand'])
        for prod in self.products:
            ws.append([prod['ID'], prod['Name'], prod['Incoming'], prod['Outgoing'], prod['OnHand']])
        wb.save(filename)

    def read_from_excel(self, filename='products.xlsx'):
        try:
            wb = load_workbook(filename)
            ws = wb.active
            self.products = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                id, name, incoming, outgoing, on_hand = row
                self.products.append({'ID': id, 'Name': name, 'Incoming': incoming, 'Outgoing': outgoing, 'OnHand': on_hand})
            print("Previous product data:")
            for prod in self.products:
                print(prod)
        except FileNotFoundError:
            print("No previous product data found.")

class Stock:
    def __init__(self):
        self.stock = []
        self.product_manager = Product()
        self.customer_manager = Customer()

    def add_stock(self, id, customer_id, product_id, qty, type):
        if any(stock['ID'] == id for stock in self.stock):
            print("ID must be unique.")
            return
        self.stock.append({'ID': id, 'CustomerID': customer_id, 'ProductID': product_id, 'Qty': qty, 'Type': type})
        self.product_manager.update_stock(product_id, type, qty)

    def add_customer(self, id, name, type, pan):
        self.customer_manager.add_customer(id, name, type, pan)

    def save_to_excel(self, filename='stock.xlsx'):
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'CustomerID', 'ProductID', 'Qty', 'Type'])
        for item in self.stock:
            ws.append([item['ID'], item['CustomerID'], item['ProductID'], item['Qty'], item['Type']])
        wb.save(filename)

    def read_from_excel(self, filename='stock.xlsx'):
        try:
            wb = load_workbook(filename)
            ws = wb.active
            self.stock = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                id, customer_id, product_id, qty, type = row
                self.stock.append({'ID': id, 'CustomerID': customer_id, 'ProductID': product_id, 'Qty': qty, 'Type': type})
            print("Previous stock data:")
            for item in self.stock:
                print(item)
        except FileNotFoundError:
            print("No previous stock data found.")

def add_new_customer():
    customer_manager = Customer()
    customer_manager.read_from_excel()
    id = input("Enter Customer ID: ")
    name = input("Enter Customer Name: ")
    type = input("Enter Customer Type (customer/vendor): ")
    pan = input("Enter PAN: ")
    customer_manager.add_customer(id, name, type, pan)
    customer_manager.save_to_excel()

def add_new_product():
    product_manager = Product()
    product_manager.read_from_excel()
    id = input("Enter Product ID: ")
    name = input("Enter Product Name: ")
    incoming = float(input("Enter Incoming Quantity: "))
    outgoing = float(input("Enter Outgoing Quantity: "))
    product_manager.add_product(id, name, incoming, outgoing)
    product_manager.save_to_excel()

def add_new_stock():
    stock_manager = Stock()
    stock_manager.product_manager.read_from_excel()
    stock_manager.customer_manager.read_from_excel()
    stock_manager.read_from_excel()

    id = input("Enter ID: ")
    customer_id = input("Enter Customer ID: ")
    product_id = input("Enter Product ID: ")
    qty = float(input("Enter Quantity: "))
    type = input("Enter Type (incoming/outgoing): ")

    stock_manager.add_stock(id, customer_id, product_id, qty, type)


    # add_customer = input("Do you want to add a new customer? (yes/no): ")
    # if add_customer.lower() == 'yes':
    #     cust_id = input("Enter Customer ID: ")
    #     name = input("Enter Customer Name: ")
    #     cust_type = input("Enter Customer Type (customer/vendor): ")
    #     pan = input("Enter PAN: ")
    #     stock_manager.add_customer(cust_id, name, cust_type, pan)

    stock_manager.save_to_excel()
    stock_manager.product_manager.save_to_excel()
    stock_manager.customer_manager.save_to_excel()

def main():
    while True:
        print("\nWhich class data do you want to add?")
        choice = input("Enter 'customer', 'product', or 'stock' (or 'exit' to quit): ")

        if choice.lower() == 'customer':
            add_new_customer()
        elif choice.lower() == 'product':
            add_new_product()
        elif choice.lower() == 'stock':
            add_new_stock()
        elif choice.lower() == 'exit':
            break
        else:
            print("Invalid choice! Please enter 'customer', 'product', 'stock', or 'exit'.")

if __name__ == "__main__":
    main()
