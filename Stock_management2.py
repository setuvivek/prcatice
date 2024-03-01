
import itertools
from openpyxl import load_workbook

ws = load_workbook('Cust_Master.xlsx')
if "CUST_MASTER" and "PRODUCT_MASTER" and "STOCK_MASTER" not in ws.sheetnames:
    ws.create_sheet("CUST_MASTER")
    ws["CUST_MASTER"].append(("CUST_ID", "CUST_NAME", "CUST_TYPE", "CUST_PAN"))
    ws.create_sheet("PRODUCT_MASTER")
    ws["PRODUCT_MASTER"].append(("PROD_ID", "PROD_NAME", 'PROD_INCOMING', 'PROD_OUT_GOING', 'PROD_STOCK'))
    ws.create_sheet("STOCK_MASTER")
    ws["STOCK_MASTER"].append(("STOCK_ID", "CUST_ID", "PROD_ID", "STOCK_QUANTITY", "STOCK_TYPE_I_O"))
    ws.save('Cust_Master.xlsx')
    cut = ws["CUST_MASTER"].max_row
    pro = ws["PRODUCT_MASTER"].max_row
    sto = ws["STOCK_MASTER"].max_row
    print(cut, pro, sto)


class Customer_Master:
    id = ws["CUST_MASTER"].max_row
    CUST_ID = itertools.count(id)
    CUST_NAME = str()
    CUST_TYPE = []
    CUST_PAN = str()
    sh1 = ws["CUST_MASTER"]
    @classmethod
    def CUST_Data_Get(self):
        while True:
            print('Customer Details')
            choice = int(input(
                "1.insert data\n2.print data\n11.ENTER 11 TO EXIT\nCHOSE YOUR OPERATION :- "))
            print()
            if choice == 1:
                cust_id_count = next(self.CUST_ID)
                print("CUSTOMER ID :- ", cust_id_count)
                self.CUST_NAME = input("CUSTOMER NAME :- ")
                if self.CUST_NAME == '':
                    print("NAME CAN'T BLANK\n")
                    return Customer_Master().CUST_Data_Get()
                print("1.customer\n2.vendor\n3.customer,vendor\nCHOSE YOUR OPERATION :- ")
                type_c = int(input("CUSTOMER TYPE *(customer, vendor) :- "))
                if type_c == 1:
                    self.CUST_TYPE = ["customer"]
                elif type_c == 2:
                    self.CUST_TYPE = ["vendor"]
                elif type_c == 3:
                    self.CUST_TYPE = ["customer", "vendor"]
                else:
                    print("ENTER VALID OPTION.\n")
                    self.id = -1
                    return Customer_Master().CUST_Data_Get()
                self.CUST_PAN = input("CUSTOMER PAN :- ")
                if self.CUST_PAN == '':
                    print("PAN NO. CAN'T BLANK\n")
                    self.id = -1
                    return Customer_Master().CUST_Data_Get()
                print()
                for k in self.sh1:
                    if self.CUST_PAN == str(k[3].value):
                        print("PAN ALREADY EXISTS\n")
                        self.id = -1
                        return Customer_Master().CUST_Data_Get()
                self.sh1.append((cust_id_count, self.CUST_NAME, ",".join(self.CUST_TYPE), self.CUST_PAN))
                ws.save('Cust_Master.xlsx')
            elif choice == 2:
                for j in self.sh1:
                    for k in j:
                        print(k.value, end=" ")
                    print()
                return Customer_Master().CUST_Data_Get()
            elif choice == 11:
                break
            else:
                print("Enter valid option")


class Product_Master(Customer_Master):
    id1 = ws["PRODUCT_MASTER"].max_row
    PROD_ID = itertools.count(id1)
    PROD_NAME = str()
    PROD_INCOMING = float()
    PROD_OUT_GOING = float()
    PROD_STOCK = float()
    sh2 = ws["PRODUCT_MASTER"]

    def PRODU_GET_DATA(self):
        while True:
            print('Product Master')
            choice = int(input(
                "1.insert data\n2.print data\n11.ENTER 11 TO EXIT\nCHOSE YOUR OPERATION :- "))
            print()
            if choice == 1:
                prod_id_count = next(self.PROD_ID)
                print("PRODUCT ID :- ", prod_id_count)
                self.PROD_NAME = input("PRODUCT NAME :- ")
                if self.PROD_NAME == '':
                    print("NAME CAN'T BLANK\n")
                    self.id1 = -1
                    return Product_Master().PRODU_GET_DATA()
                for a in self.sh2:
                    if self.PROD_NAME == a[1].value:
                        print("NAME ALREADY EXISTS")
                        self.id1 = -1
                        return Product_Master().PRODU_GET_DATA()
                self.sh2.append(
                    (prod_id_count, self.PROD_NAME, self.PROD_INCOMING, self.PROD_OUT_GOING, self.PROD_STOCK))
                ws.save('Cust_Master.xlsx')
            elif choice == 2:
                for j in self.sh2:
                    for k in j:
                        print(k.value, end=" ")
                    print()
                return Product_Master().PRODU_GET_DATA()
            elif choice == 11:
                break
            else:
                print("Enter valid option")


class Stock_Master(Product_Master, Customer_Master):
    id2 = ws["PRODUCT_MASTER"].max_row
    STOCK_ID = itertools.count(id2)
    PROD_ID = 0
    CUST_ID = 0
    STOCK_QUANTITY = float()
    STOCK_TYPE_I_O = str()
    State = str()
    sh3 = ws["STOCK_MASTER"]

    def Stock_Data_GET(self):
        while True:
            print('Stock Master')
            choice = int(input(
                "1.insert data\n2.print data\n11.ENTER 11 TO EXIT\nCHOSE YOUR OPERATION :- "))
            print()
            if choice == 1:
                stock_id_count = next(self.STOCK_ID)
                print("STOCK INCOMING OUTGOING ID :- ", stock_id_count)
                self.CUST_ID = int(input("ENTER CUSTOMER ID :- "))
                lis_cust_id = []
                for e in self.sh1:
                    lis_cust_id.append(e[0].value)
                if self.CUST_ID not in lis_cust_id:
                    print('PLEASE ENTER VALID CUSTOMER ID.\n')
                    work = int(
                        input("you want to add that customer ? \n1.YES\n2.NO\nCHOSE YOUR OPERATION :- "))
                    if work == 1:
                        self.id2 = -1
                        return Customer_Master().CUST_Data_Get()
                    self.id2 = -1
                    return Stock_Master().Stock_Data_GET()

                self.PROD_ID = int(input("ENTER PRODUCT ID :- "))
                lis_prod_id = []
                for j in self.sh2:
                    lis_prod_id.append(j[0].value)
                if self.PROD_ID not in lis_prod_id:
                    print('PLEASE ENTER VALID PRODUCT ID.\n')
                    work = int(
                        input("you want to add that product ? \n1.YES\n2.NO\nCHOSE YOUR OPERATION :- "))
                    if work == 1:
                        self.id2 = -1
                        return Product_Master().PRODU_GET_DATA()
                    self.id2 = -1
                    return Stock_Master().Stock_Data_GET()
                self.STOCK_QUANTITY = float(input("ENTER THE QUANTITY THAT YOU WANT TO IN OR OUT :- "))
                choice_prod_i_o = int(
                    input("1.PRODUCT INCOMING\n2.PRODUCT OUTGOING\nCHOSE YOUR OPERATION :- "))
                if choice_prod_i_o == 1:
                    for k in self.sh2:
                        if self.PROD_ID == k[0].value:
                            k[2].value += self.STOCK_QUANTITY
                            k[4].value += self.STOCK_QUANTITY
                    self.STOCK_TYPE_I_O = "PRODUCT INCOMING"
                elif choice_prod_i_o == 2:
                    for j in self.sh2:
                        if isinstance(j[-1].value, str):
                            continue
                        if self.PROD_ID == j[0].value:
                            if self.STOCK_QUANTITY > int(j[-1].value):
                                print('PLEASE CHECK STOCK.\n')
                                self.id2 = -1
                                return Stock_Master().Stock_Data_GET()
                            j[3].value += self.STOCK_QUANTITY
                            j[4].value -= self.STOCK_QUANTITY
                        self.STOCK_TYPE_I_O = "PRODUCT OUTGOING"
                else:
                    self.id2 = -1
                    print("Enter valid option")
                self.sh3.append(
                    (stock_id_count, self.CUST_ID, self.PROD_ID, self.STOCK_QUANTITY, self.STOCK_TYPE_I_O))
                ws.save('Cust_Master.xlsx')

            elif choice == 2:
                for k in self.sh3:
                    for j in k:
                        print(j.value, end=" ")
                    print()
                return Stock_Master().Stock_Data_GET()
            elif choice == 11:
                break
            else:
                print("Enter valid option")


while True:
    try:
        i = int(input(
            "1.Customer_Master\n2.Product_Master\n3.Stock_Move_Master\n11.ENTER 11 TO EXIT\nCHOSE YOUR OPERATION :- "))
        if i == 1:
            print()
            Customer_Master().CUST_Data_Get()
        elif i == 2:
            Product_Master().PRODU_GET_DATA()
            print()
        elif i == 3:
            print()
            Stock_Master().Stock_Data_GET()
            ws.close()
        elif i == 11:
            break
        else:
            print("Enter Valid Operation.")
    except Exception as error:
        c = Customer_Master()
        c.id = -1
        p = Product_Master()
        p.id1 = -1
        s = Stock_Master()
        s.id2 = -1
        ws.close()
        print("you Interrupt the System.", type(error).__name__, "–", error)
        exit()
    except:
        c = Customer_Master()
        c.id = -1
        p = Product_Master()
        p.id1 = -1
        s = Stock_Master()
        s.id2 = -1
        ws.close()
        print("you Interrupt the System.")
        exit()
print("THANK YOU......")
