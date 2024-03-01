import itertools

import openpyxl as ol

wb = ol.load_workbook('stock_manage.xlsx')
for sheet in wb:
    if sheet.max_row == 1 and sheet.max_column == 1:
        wb.remove(sheet)
    if 'customer' and 'product' and 'stock' not in wb.sheetnames:
        wb.create_sheet("customer")
        wb["customer"].append(("cust_id", "cust_name", "cust_type", "cust_pan"))
        wb.create_sheet('product')
        wb["product"].append(("prod_id", "prod_name", 'prod_in', 'prod_out', 'onhand'))
        wb.create_sheet('stock')
        wb["stock"].append(("id", "cust_id", "prod_id", "stock_qty", "stock_in_out"))
        wb.save('stock_manage.xlsx')
        cut = wb["customer"].max_row
        pro = wb["product"].max_row
        sto = wb["stock"].max_row


class customer_master:
    cust_info = {}
    cust_name = str()
    cust_type = []
    cust_pan = int()
    id = wb["customer"].max_row
    cust_id = itertools.count(id)
    sheet1 = wb["customer"]

    def __init__(self):
        self.load_data()

    def load_data(self):
        sheet1 = wb["customer"]
        for row in sheet1.iter_rows(min_row=2, values_only=True):
            cust_id, cust_name, cust_types, cust_pan = row
            self.cust_info[cust_id] = {'cust_name': cust_name, 'cust_type': cust_types, 'cust_pan': cust_pan}

    def cust_data(self):
        while True:
            print('customer details')

            option = int(input('1.insert_data\n2.print_data\n3.press 3 to exit\nwhat do you want to do : '))

            if option == 1:
                cust_count = next(self.cust_id)
                print('cust_id :', cust_count)

                self.cust_name = input('cust_name :')
                print('1.customer\n2.vendor\n3.customer,vendor')
                cust_type_option = int(input('enter cust_type : '))
                if cust_type_option == 1:
                    self.cust_type = ['customer']
                elif cust_type_option == 2:
                    self.cust_type = ['vendor']
                elif cust_type_option == 3:
                    self.cust_type = ['customer', 'vendor']
                else:
                    print('enter valid cust_type')
                    return customer_master().cust_data()

                self.cust_pan = input('enter pan no:')
                if self.cust_pan == '':
                    print('pan no. is mandatory please enter pan no..')
                    return customer_master()

                for x in range(1, len(self.cust_info) + 1):
                    if self.cust_pan == self.cust_info[x].get("cust_pan"):
                        print("please check pan no..\n")
                        return customer_master()
                self.cust_info[cust_count] = {'cust_name': self.cust_name,
                                              'cust_type': self.cust_type, 'cust_pan': self.cust_pan}
                sheets = wb.worksheets[0]
                pan = []
                for row in sheets:
                    pans = row[3].value

                    pan.append(pans)

                # print(pan)
                if self.cust_pan in pan:
                    print('pan is already exists in data please check')
                    return customer_master().cust_data()
                else:
                    self.sheet1.append((cust_count, self.cust_name, ",".join(self.cust_type), self.cust_pan))
                    wb.save('stock_manage.xlsx')
                    return customer_master().cust_data()
            elif option == 2:
                print(self.cust_info)
                return customer_master().cust_data()

            elif option == 3:
                print('thank you visit again')
                break
            else:
                print('enter valid option')
                return customer_master().cust_data()


# c1 = customer_master()
# c1.cust_data()


class product_master(customer_master):
    prod_info = {}
    prod_name = str()
    prod_in = float()
    prod_out = float()
    prod_on_hand = 0.0
    id1 = wb["product"].max_row
    prod_id = itertools.count(id1)
    sheet2 = wb["product"]

    def __init__(self):
        super().__init__()
        self.load_data()

    def load_data(self):
        super().load_data()
        sheet2 = wb["product"]
        for row in sheet2.iter_rows(min_row=2, values_only=True):
            prod_id, prod_name, prod_in, prod_out, prod_on_hand = row
            self.prod_info[prod_id] = {'prod_name': prod_name, 'prod_in': prod_in, 'prod_out': prod_out, 'prod_on_hand': prod_on_hand}

    def prod_data(self):
        while True:
            print('product details')

            option = int(input('1.insert_data\n2.print_data\n3.press 3 to exit\nwhat do you want to do : '))
            if option == 1:
                prod_count = next(self.prod_id)
                print('prod_id :', prod_count)

                self.prod_name = input('prod_name : ')
                if self.prod_name == '':
                    print('product name is mandatory')
                    return product_master()

                for x in range(1, len(self.prod_info) + 1):
                    if self.prod_name.casefold() == self.prod_info[x].get('prod_name'):
                        print('This product is  already there you can not enter it again.. ')
                        return product_master()
                self.prod_info[prod_count] = {'prod_name': self.prod_name,
                                              'prod_in': 0, 'prod_out': 0,
                                              'prod_on_hand': self.prod_on_hand}
                sheets = wb.worksheets[1]
                prod_names = []
                for row in sheets:
                    name = row[1].value
                    prod_names.append(name)
                    prod_names = [x.lower() for x in prod_names]
                    # print(product_names)

                if self.prod_name in prod_names:
                    print('product is already exists in data please check')
                    return product_master().prod_data()
                else:
                    self.sheet2.append((prod_count, self.prod_name, self.prod_in, self.prod_out, self.prod_on_hand))
                    wb.save('stock_manage.xlsx')

            elif option == 2:
                print(self.prod_info)
                return product_master().prod_data()

            elif option == 3:
                print('thank you visit again')
                break
            else:
                print('enter valid option')
                return product_master().prod_data()


# p1 = product_master()
# p1.prod_data()

class stock_move(product_master, customer_master):
    stock_info = {}
    prod_id = 0
    cust_id = 0
    stock_qty = 0
    stock_in_out = str()
    state = str()
    id2 = wb["stock"].max_row
    stock_id = itertools.count(id2)
    sheet3 = wb["stock"]

    def __init__(self):
        super().__init__()
        self.load_data()

    def load_data(self):
        super().load_data()
        sheet3 = wb["stock"]
        for row in sheet3.iter_rows(min_row=2, values_only=True):
            stock_id, cust_id, prod_id, stock_qty, stock_in_out = row
            self.stock_info[stock_id] = {'cust_id': cust_id, 'prod_id': prod_id, 'stock_qty': stock_qty, 'stock_in_out': stock_in_out}

    def update_product_sheet(self):
        for prod_id, details in self.prod_info.items():
            for x in self.sheet2.iter_rows():
                if x[0].value == prod_id:
                    x[2].value = details['prod_in']
                    x[3].value = details['prod_out']
                    x[4].value = details['prod_on_hand']
                    break

    def stock_data(self):
        while True:
            print('stock details')

            option = int(input('1.insert_data\n2.print_data\n3.press 3 to exit\nwhat do you want to do : '))

            if option == 1:
                stock_count = next(self.stock_id)
                print('stock_id :', stock_count)

                self.cust_id = int(input("enter cust_id : "))
                if self.cust_id not in super().cust_info.keys():
                    print('please enter valid cust_id :')
                    option = int(input("you want to add that customer ? \n1.yes\n2.no\nwhat do you want to do : "))
                    if option == 1:
                        return customer_master().cust_data()
                    return stock_move().stock_data()

                self.prod_id = int(input('enter prod_id : '))
                if self.prod_id not in super().prod_info.keys():
                    print('please enter valid prod_id :')
                    option = int(input("you want to add that product ? \n1.yes\n2.no\nwhat do you want to do : "))
                    if option == 1:
                        return product_master().prod_data()
                    return stock_move().stock_data()

                self.stock_qty = float(input('enter qty : '))
                if self.stock_qty < 1:
                    print('please enter proper qty !..')
                    return stock_move().stock_data()

                option_prod_in_out = int(input("1.prod_in\n2.prod_out\nwhat do you want to do :"))
                if option_prod_in_out == 1:
                    sum_of_in = self.stock_qty + super().prod_info[self.prod_id].get(
                        "prod_in")
                    sum_of_stock = self.stock_qty + super().prod_info[self.prod_id].get("prod_on_hand")
                    super().prod_info[self.prod_id].update(
                        {"prod_in": sum_of_in, "prod_on_hand": sum_of_stock})
                    self.stock_in_out = "prod_in"
                elif option_prod_in_out == 2:
                    if self.stock_qty > super().prod_info[self.prod_id].get("prod_on_hand"):
                        print('please check you do not have enough items.\n')
                        return stock_move().stock_data()
                    sum_of_out = self.stock_qty + super().prod_info[self.prod_id].get(
                        "prod_out")
                    sum_of_stock = super().prod_info[self.prod_id].get(
                        "prod_on_hand") - self.stock_qty
                    super().prod_info[self.prod_id].update(
                        {"prod_out": sum_of_out, "prod_on_hand": sum_of_stock})
                    self.stock_in_out = "prod_out"
                else:
                    print("Enter valid option")

                self.stock_info[stock_count] = {'prod_id': self.prod_id,
                                                'prod_qty': self.stock_qty,
                                                'stock_in_out': self.stock_in_out}

                self.sheet3.append(
                    (stock_count, self.cust_id, self.prod_id, self.stock_qty, self.stock_in_out))

                self.update_product_sheet()
                wb.save('stock_manage.xlsx')

            elif option == 2:
                print(self.stock_info)
                return stock_move().stock_data()

            elif option == 3:
                print('thank you visit again ..')
                break


while True:
    i = int(input(
        '1.customer_Master\n2.product_Master\n3.Stock_Move\n4.exit\nwhat do you want to do : '))
    print()
    if i == 1:
        customer_master().cust_data()
    elif i == 2:
        product_master().prod_data()
    elif i == 3:
        stock_move().stock_data()
    else:
        print('cust_data :')
        print(customer_master().cust_info)

        print('prod_data :')
        print(product_master().prod_info)

        print('stock_data : ')
        print(stock_move().stock_info)
        break
